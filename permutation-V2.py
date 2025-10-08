from openpyxl import load_workbook


def fact(n, stringList):
    if n == 1:
        return stringList[0]
    else:
        return stringList[n - 1] + " " + fact(n - 1, stringList[:n - 1])


def permutation(text, fp):
    originalText = text.split(" ")
    textLen = len(originalText)
    original_result = []
    for j in range(textLen):
        # print("test:{}".format(test))
        tmpText = originalText[j:]
        # print("j:{}".format(j))
        # print(tmpText)
        tmpText.reverse()
        # print(tmpText)
        
        for i in range(len(tmpText)):
            #print(tmpText)
            #print("i:{}".format(i))
            temp_result = fact(len(tmpText[len(tmpText) - i - 1:len(tmpText)]), tmpText[len(tmpText) - i - 1:len(tmpText)])
            #input(temp_result)
            original_result.append(temp_result)
            #input(result)
        #fp.write(temp_result)
    #print(original_result)
    #print(len(original_result))
    unique_result = []
    [unique_result.append(x) for x in original_result if x not in unique_result]
    fp.write(','.join(unique_result))
    fp.write("\n")
    result = [f"FM:({x})" for x in unique_result]
    uspto_str = "(" + ' '.join(result) + ") AND LD:true"
    fp.write(uspto_str)
    fp.write("\n")


if __name__ == '__main__':
    # test = ["Vet at Heart Dog Doctor Veterinarian Nurse", "Ask Me About My Foster Dog"]
    file = input("\n文件路径：\n")
    wb = load_workbook(file.replace("\"", ""))
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    print(ws.title)
    print(ws.max_row)
    print(ws.max_column)
    keyIn = input("按回车:开始\n")
    if '/' in keyIn:
        path = keyIn[1:] + ".csv"
        print(path)
    else:
        path = file.replace("\"", "").split('.')[0] + ".csv"
        print(path)
    text = []
    for i in range(1, ws.max_row+1):
        text.append(ws.cell(i, 1).value)
    fp = open(path, 'a')
    for i in range(len(text)):
        fp.write(text[i] + ',')
        permutation(text[i], fp)
    fp.close()
    input("搞定,按回车结束")